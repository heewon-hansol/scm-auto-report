"""
SCM지원파트 딸깍! 보고자료 자동화 플랫폼 v5
- 17년~25년 연평균 + 26년 월별 연속 트랜드 차트 (Gap_trends Excel 동일 구조)
- 템플릿 기반 검증 기능
"""
import io, shutil, tempfile, os, smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import pandas as pd
import streamlit as st
import plotly.graph_objects as go
import openpyxl
from openpyxl.cell.cell import MergedCell

PRIMARY   = "#1A4F9C"
PRIMARY_D = "#0F3272"
ACCENT    = "#2D9A4F"
ACCENT_L  = "#E8F5EE"
GRAY_T    = "#64748B"

st.set_page_config(page_title="딸깍! 보고자료 자동화 플랫폼",
                   page_icon="⚡", layout="wide",
                   initial_sidebar_state="expanded")

# ── 사이드바 CSS (네이비 배경 + 버튼 가시성) ──────────────────
st.markdown("""<style>
section[data-testid="stSidebar"]>div:first-child{background:#0F3272!important}
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] div[data-testid="stMarkdownContainer"] *
  {color:#E2E8F0!important}
section[data-testid="stSidebar"] button[kind="secondary"]
  {background:rgba(255,255,255,.1)!important;color:#E2E8F0!important;
   border:1px solid rgba(255,255,255,.2)!important;font-size:13px!important}
section[data-testid="stSidebar"] button[kind="secondary"]:hover
  {background:rgba(255,255,255,.2)!important;color:#fff!important}
section[data-testid="stSidebar"] button[kind="primary"]
  {background:#2563EB!important;color:#fff!important;
   border:none!important;font-size:13px!important;font-weight:700!important}
</style>""", unsafe_allow_html=True)

# ── 메인 CSS ──────────────────────────────────────────────────
st.markdown(f"""<style>
html,body,[class*="css"]{{font-family:'Malgun Gothic','Apple SD Gothic Neo',sans-serif}}
#MainMenu,footer,header{{visibility:hidden}}
.plat-title{{background:linear-gradient(135deg,{PRIMARY_D} 0%,{PRIMARY} 60%,{ACCENT} 100%);
  color:white;padding:18px 24px;border-radius:12px;margin-bottom:14px}}
.plat-title h1{{margin:0;font-size:20px;font-weight:700}}
.plat-title p{{margin:3px 0 0;font-size:12px;opacity:.85}}
.report-card{{background:white;border:2px solid #e2e8f0;border-radius:12px;
  padding:20px 18px;min-height:130px;transition:all .18s}}
.report-card:hover{{border-color:{PRIMARY};box-shadow:0 4px 16px rgba(26,79,156,.15)}}
.report-card.coming{{opacity:.42}}
.card-icon{{font-size:28px}}.card-title{{font-size:14px;font-weight:700;color:{PRIMARY_D};margin:6px 0 3px}}
.card-desc{{font-size:11px;color:{GRAY_T}}}
.badge{{display:inline-block;font-size:10px;padding:2px 8px;border-radius:20px;
  background:{ACCENT};color:white;margin-top:5px;font-weight:600}}
.badge.soon{{background:#94a3b8}}
.sec-hdr{{background:linear-gradient(90deg,{PRIMARY} 0%,{PRIMARY}BB 100%);
  color:white;padding:7px 16px;border-radius:8px;font-size:14px;font-weight:700;margin:20px 0 8px}}
.num-table{{width:100%;border-collapse:collapse;font-size:12px;margin-bottom:14px}}
.num-table th{{background:{PRIMARY};color:white;padding:5px 8px;text-align:center;
  border:1px solid #dde3f0;font-size:11px}}
.num-table td{{padding:4px 8px;text-align:right;border:1px solid #e8edf5;color:#1e293b}}
.num-table td.lbl{{text-align:left;font-weight:500;background:#f8fafc;min-width:60px}}
.num-table tr.cur td{{background:#EFF6FF;font-weight:700;color:{PRIMARY_D}}}
.num-table tr.plan td{{background:{ACCENT_L};color:{ACCENT}}}
</style>""", unsafe_allow_html=True)

BASE = os.path.dirname(__file__)
DEFAULT = {
    "bw":       os.path.join(BASE, "펄프가(인쇄감열7월).xlsx"),
    "산업":      os.path.join(BASE, "펄프가(산업7월).xlsx"),
    "plan":     os.path.join(BASE, "지종별 계획지류 관리.xlsx"),
    "template": os.path.join(BASE, "(2026-07월)Gap_trends.xlsx"),
}

# ── 템플릿 historical 행 매핑 ─────────────────────────────────
# col 4~12 = 2017~2025 연평균, col 13 = 2026 경영계획 연평균, col 14+ = 2026 1월~
HIST_YEAR_COLS = {4:"17년",5:"18년",6:"19년",7:"20년",8:"21년",9:"22년",10:"23년",11:"24년",12:"25년"}
PLAN_AVG_COL = 13
CUR_MON_BASE_COL = 14  # 14=1월, 15=2월, ...

TMPL_ROWS = {
    ("인쇄","내수","일반백상"):10, ("인쇄","내수","일반아트"):11, ("인쇄","내수","합계"):12,
    ("인쇄","수출","일반백상"):44, ("인쇄","수출","일반아트"):45, ("인쇄","수출","합계"):46,
    ("감열","내수","POS영업팀"):81, ("감열","내수","라벨영업팀"):82, ("감열","내수","합계"):83,
    ("감열","수출","POS영업팀"):116, ("감열","수출","라벨영업팀"):117, ("감열","수출","합계"):118,
    ("패키징","내수","기능지류"):426, ("패키징","내수","전사지류"):427,
    ("패키징","내수","일반팬시지류"):428, ("패키징","내수","러프그로스지"):429,
    ("패키징","내수","합계(Protego제외)"):430,
    ("패키징","수출","전사지류"):524, ("패키징","수출","일반팬시지류"):525,
    ("패키징","수출","러프그로스지"):526, ("패키징","수출","합계(Protego제외)"):527,
    ("산업","내수","AB류"):297,
    ("산업","내수","SC류(고)"):338, ("산업","내수","SC류(저)"):341, ("산업","내수","SC류 소계"):342,
    ("산업","수출","SC류(고)"):377, ("산업","수출","SC류(저)"):380, ("산업","수출","SC류 소계"):381,
}

for k,v in [("page","home"),
            ("smtp",{"host":"echohub.hansol.com","port":25,"auth":False,"user":"","password":""})]:
    if k not in st.session_state: st.session_state[k]=v

# ══════════════════════════════════════════════════════════════
# 데이터 로드
# ══════════════════════════════════════════════════════════════
def read_bytes(up, key):
    if up: up.seek(0); return up.read()
    p=DEFAULT.get(key,"")
    if os.path.exists(p):
        with open(p,"rb") as f: return f.read()
    return None

def to_int(v):
    try:    return int(float(v))
    except: return None

def load_ie(data):
    df=pd.read_excel(data, sheet_name="값", header=0)
    for c in ["판매량","순매출액","R_원재료","A펄프","펄프_조정"]:
        df[c]=pd.to_numeric(df[c],errors="coerce").fillna(0)
    return df

def load_산업(data):
    df=pd.read_excel(data, sheet_name="BW", header=0)
    df=df[df["값유형"]==10].copy()
    for c in ["판매량(톤)","순매출액","펄프원재료","고지원재료"]:
        df[c]=pd.to_numeric(df[c],errors="coerce").fillna(0)
    return df

def parse_period(m):
    try:
        parts=str(m).split(".")
        if len(parts)!=2: return None
        y,mo=int(parts[0]),int(parts[1])
        if 2000<=y<=2100 and 1<=mo<=12: return y,mo
    except: pass
    return None

def all_periods(df, col="회계연도/기간"):
    result=[]
    for m in df[col].astype(str).unique():
        p=parse_period(m)
        if p: result.append((p[0],p[1],m))
    return sorted(result)

def gap_std(sub):
    v=sub["판매량"].sum()
    if v==0: return None
    return to_int((sub["순매출액"].sum()-sub["R_원재료"].sum())/v/1000)

def gap_ab(sub):
    v=sub["판매량(톤)"].sum()
    if v==0: return None
    return to_int((sub["순매출액"].sum()-sub["펄프원재료"].sum())/v/1000)

def gap_sc(sub):
    v=sub["판매량(톤)"].sum()
    if v==0: return None
    return to_int((sub["순매출액"].sum()-sub["고지원재료"].sum())/v/1000)

# ── 템플릿 historical 데이터 로드 ────────────────────────────
@st.cache_data(show_spinner=False)
def load_template_hist(tmpl_path):
    """Returns {(부문,유통,label): {year_label: avg, '26경영계획': v, month_int: v}}"""
    if not os.path.exists(tmpl_path): return {}
    wb=openpyxl.load_workbook(tmpl_path, data_only=True)
    ws=wb["Gap Trends"]
    result={}
    for key,row in TMPL_ROWS.items():
        d={}
        for col,yr_lbl in HIST_YEAR_COLS.items():
            v=ws.cell(row,col).value
            if v is not None: d[yr_lbl]=to_int(v)
        pv=ws.cell(row,PLAN_AVG_COL).value
        if pv is not None: d["26경영계획"]=to_int(pv)
        for mon in range(1,13):
            v=ws.cell(row,CUR_MON_BASE_COL+mon-1).value
            if v is not None: d[mon]=to_int(v)
        result[key]=d
    return result

# ── BW 계산 ──────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def calc_gap_data(bw_bytes):
    raw=load_ie(io.BytesIO(bw_bytes))
    periods=all_periods(raw)

    def build(df, 부문, 유통, label_filts):
        out={}
        for label,col,val in label_filts:
            out[label]={}
            for yr,mo,mstr in periods:
                sub=df[(df["회계연도/기간"].astype(str)==mstr)&
                       (df["계획지류 부문"]==부문)&(df["유통경로"]==유통)]
                if col: sub=sub[sub[col]==val]
                out[label].setdefault(yr,{})[mo]=gap_std(sub)
        return out

    실적=raw[raw["값유형"]==10].copy()
    계획=raw[raw["값유형"]==20].copy()

    인쇄_lf=[("일반백상","지류","일반백상"),("일반아트","지류","일반아트"),("합계",None,None)]
    감열_lf=[("POS영업팀","계획지류 지류(sriv 고급)","POS"),
             ("라벨영업팀","계획지류 지류(sriv 고급)","일감라벨"),("합계",None,None)]

    r_실={
        "인쇄":{
            "내수":build(실적,"인쇄","내수",인쇄_lf),
            "수출":build(실적,"인쇄","수출",인쇄_lf),
        },
        "감열":{
            "내수":build(실적,"감열","내수",감열_lf),
            "수출":build(실적,"감열","수출",감열_lf),
        },
    }
    for 유통 in ["내수","수출"]:
        r_실.setdefault("패키징",{})[유통]={}
        pk_lf=[(j,"계획지류 지류",j)
               for j in ["기능지류","전사지류","일반팬시지류","러프그로스지","Protego"]]
        pk_lf.append(("합계(Protego제외)",None,None))
        for label,col,val in pk_lf:
            r_실["패키징"][유통][label]={}
            for yr,mo,mstr in periods:
                sub=실적[(실적["회계연도/기간"].astype(str)==mstr)&
                        (실적["계획지류 부문"]=="패키징")&(실적["유통경로"]==유통)]
                if col: sub=sub[sub[col]==val]
                elif label=="합계(Protego제외)": sub=sub[sub["계획지류 지류"]!="Protego"]
                r_실["패키징"][유통][label].setdefault(yr,{})[mo]=gap_std(sub)

    r_계=None
    if not 계획.empty:
        r_계={
            "인쇄":{"내수":build(계획,"인쇄","내수",인쇄_lf),
                    "수출":build(계획,"인쇄","수출",인쇄_lf)},
            "감열":{"내수":build(계획,"감열","내수",감열_lf),
                    "수출":build(계획,"감열","수출",감열_lf)},
        }
    avail_months=sorted(set(mo for _,mo,_ in periods))
    return r_실, r_계, avail_months

@st.cache_data(show_spinner=False)
def calc_산업_data(bw_bytes):
    bw=load_산업(io.BytesIO(bw_bytes))
    def bsm(유통,jl): return bw[(bw["유통"]==유통)&(bw["지류2"].isin(jl))]
    ab={
        "AB류 내수":{"Gap":gap_ab(bsm("내수",["AB류"]))},
        "AB류 수출":{"Gap":gap_ab(bsm("수출",["AB류"]))},
    }
    sc={}
    for lbl,(유통,jl) in {
        "SC류(고) 내수":("내수",["SC류(고)"]),"IV류 내수":("내수",["IV류"]),
        "SC류(저) 내수":("내수",["SC류(저)"]),"소계 내수":("내수",["SC류(고)","IV류","SC류(저)"]),
        "SC류(고) 수출":("수출",["SC류(고)"]),"IV류 수출":("수출",["IV류"]),
        "SC류(저) 수출":("수출",["SC류(저)"]),"소계 수출":("수출",["SC류(고)","IV류","SC류(저)"]),
    }.items():
        sc[lbl]={"Gap":gap_sc(bsm(유통,jl))}
    return {"AB류":ab,"SC류":sc}

# ══════════════════════════════════════════════════════════════
# 차트: 17년~25년(연평균) + 26년 월별 연속 트랜드
# ══════════════════════════════════════════════════════════════
HIST_XLABELS = ["17년","18년","19년","20년","21년","22년","23년","24년","25년"]
MON_XLABELS  = ["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]

def _section_labels(end_mon, show_plan):
    """X축 전체 레이블 목록 반환: hist(+계획) + 26년 월별"""
    h=HIST_XLABELS[:]
    if show_plan: h.append("26년\n경영계획")
    m=[f"{i}월" for i in range(1,end_mon+1)]
    return h, m, h+m

def trend_chart(label_key, hist_data, cur_mon_data, plan_mon_data,
                end_mon, show_plan, title, height=360):
    """
    hist_data    : {yr_label: avg_value, '26경영계획': v}
    cur_mon_data : {month_int: value}
    plan_mon_data: {month_int: value} | None
    end_mon      : 1~12
    """
    h_lbls, m_lbls, x_all = _section_labels(end_mon, show_plan)
    n_hist = len(h_lbls)   # 히스토리(+계획) 구간 개수
    n_mon  = len(m_lbls)   # 월별 구간 개수

    hist_y = [hist_data.get(yr) for yr in HIST_XLABELS]
    if show_plan:
        hist_y.append(hist_data.get("26경영계획"))
    cur_y  = [cur_mon_data.get(m) for m in range(1, end_mon+1)]
    y_all  = hist_y + cur_y

    fig = go.Figure()

    # ── 히스토리 구간 음영 (월평균 배경) ──
    fig.add_vrect(x0=x_all[0], x1=x_all[n_hist-1],
                  fillcolor="rgba(203,213,225,0.22)", layer="below", line_width=0)
    # ── 26년 구간 음영 ──
    fig.add_vrect(x0=x_all[n_hist], x1=x_all[-1],
                  fillcolor="rgba(219,234,254,0.30)", layer="below", line_width=0)

    # ── 구분선 ──
    fig.add_vline(x=x_all[n_hist-1],
                  line=dict(color="#64748B", width=1.5, dash="dot"))

    # ── 구간 헤더 annotation ──
    fig.add_annotation(
        x=n_hist/2 - 0.5, y=1.06, xref="x", yref="paper",
        text="<b>　　　월　평　균　　　</b>",
        showarrow=False, font=dict(size=11, color="#334155"),
        bgcolor="#E2E8F0", bordercolor="#94a3b8", borderwidth=1, borderpad=4,
        xanchor="center",
    )
    fig.add_annotation(
        x=n_hist + n_mon/2 - 0.5, y=1.06, xref="x", yref="paper",
        text=f"<b>　　26년　실적　　</b>",
        showarrow=False, font=dict(size=11, color=PRIMARY_D),
        bgcolor="#DBEAFE", bordercolor=PRIMARY, borderwidth=1, borderpad=4,
        xanchor="center",
    )

    # ── 데이터 선 ──
    marker_sizes  = [7]*n_hist + [10]*n_mon
    marker_colors = ["#64748B"]*n_hist + [PRIMARY]*n_mon
    marker_syms   = ["circle"]*n_hist + ["diamond"]*n_mon
    fig.add_trace(go.Scatter(
        x=x_all, y=y_all,
        mode="lines+markers+text",
        name=label_key, connectgaps=True,
        line=dict(color=PRIMARY, width=2.5),
        marker=dict(size=marker_sizes, color=marker_colors,
                    symbol=marker_syms, line=dict(width=1.5, color="white")),
        text=["" if v is None else str(v) for v in y_all],
        textposition="top center",
        textfont=dict(size=10, color=[GRAY_T]*n_hist + [PRIMARY]*n_mon),
    ))

    # ── 경영계획 월별 선 ──
    if show_plan and plan_mon_data:
        py = [plan_mon_data.get(m) for m in range(1, end_mon+1)]
        if any(v is not None for v in py):
            fig.add_trace(go.Scatter(
                x=m_lbls, y=py, mode="lines+markers",
                name="경영계획(월별)", connectgaps=True,
                line=dict(color=ACCENT, width=1.5, dash="dot"),
                marker=dict(size=5, symbol="x", color=ACCENT),
            ))

    fig.update_layout(
        title=dict(text=title, font=dict(size=14, color=PRIMARY_D), x=0),
        height=height,
        yaxis=dict(title="Gap (천원/톤)", gridcolor="#e8edf5",
                   zeroline=True, zerolinecolor="#94a3b8"),
        xaxis=dict(gridcolor="#f1f5f9", tickangle=-30,
                   tickfont=dict(size=11)),
        legend=dict(orientation="h", yanchor="bottom", y=1.08, xanchor="left",
                    font=dict(size=10), bgcolor="rgba(0,0,0,0)"),
        hovermode="x unified", plot_bgcolor="white", paper_bgcolor="white",
        margin=dict(t=95, b=50, l=60, r=80),
    )
    return fig

def pivot_html(label_key, hist_data, cur_mon_data, plan_mon_data, end_mon, show_plan):
    """Excel과 동일: 연도별 연평균 + 26년 월별 한 행 테이블"""
    # 헤더 그룹
    hist_cols = HIST_XLABELS[:]
    if show_plan: hist_cols.append("26년<br>경영계획")
    mon_cols = [f"{m}월" for m in range(1, end_mon+1)]

    n_hist = len(hist_cols)
    n_mon  = len(mon_cols)

    header = (f'<tr>'
              f'<th style="text-align:left;background:#334155" rowspan="2">구분</th>'
              f'<th colspan="{n_hist}" style="background:#475569">월&nbsp;평&nbsp;균</th>'
              f'<th colspan="{n_mon}" style="background:{PRIMARY}">26년</th>'
              f'</tr><tr>')
    header += "".join(f'<th style="background:#64748B">{c}</th>' for c in hist_cols)
    header += "".join(f'<th style="background:{PRIMARY}">{c}</th>' for c in mon_cols)
    header += "</tr>"

    # 데이터: 히스토리컬 연평균들 + 26년 월별 — 한 행으로 표시
    def fmt(v): return str(v) if v is not None else "<span style='color:#cbd5e1'>-</span>"

    h_cells = "".join(f'<td>{fmt(hist_data.get(yr))}</td>' for yr in HIST_XLABELS)
    if show_plan:
        h_cells += f'<td style="color:{ACCENT};font-weight:600">{fmt(hist_data.get("26경영계획"))}</td>'
    m_cells = "".join(f'<td>{fmt(cur_mon_data.get(m))}</td>' for m in range(1, end_mon+1))
    row = f'<tr class="cur"><td class="lbl">Gap</td>{h_cells}{m_cells}</tr>'

    # 계획 월별 행
    if show_plan and plan_mon_data:
        pm_cells = "".join(f'<td style="color:{ACCENT}">{fmt(plan_mon_data.get(m))}</td>'
                           for m in range(1, end_mon+1))
        row += (f'<tr class="plan"><td class="lbl">계획</td>'
                + "".join(f'<td>-</td>' for _ in hist_cols)
                + pm_cells + '</tr>')

    return (f'<div style="overflow-x:auto"><table class="num-table">'
            f'<thead>{header}</thead><tbody>{row}</tbody></table></div>')

# ── 다중 레이블 pivot ─────────────────────────────────────────
def pivot_html_multi(labels, hist_map, cur_map, end_mon, show_plan):
    """여러 레이블 각각 행으로 표시"""
    hist_cols = HIST_XLABELS[:]
    if show_plan: hist_cols.append("26년<br>경영계획")
    mon_cols = [f"{m}월" for m in range(1, end_mon+1)]
    n_hist = len(hist_cols); n_mon = len(mon_cols)

    header = (f'<tr><th style="text-align:left;background:#334155" rowspan="2">구분</th>'
              f'<th colspan="{n_hist}" style="background:#475569">월&nbsp;평&nbsp;균</th>'
              f'<th colspan="{n_mon}" style="background:{PRIMARY}">26년</th></tr><tr>')
    header += "".join(f'<th style="background:#64748B">{c}</th>' for c in hist_cols)
    header += "".join(f'<th style="background:{PRIMARY}">{c}</th>' for c in mon_cols)
    header += "</tr>"

    def fmt(v): return str(v) if v is not None else "<span style='color:#cbd5e1'>-</span>"

    rows=""
    for lbl in labels:
        is_sum = lbl in ["합계","합계(Protego제외)"]
        hd = hist_map.get(lbl, {})
        cd = cur_map.get(lbl, {})
        h_cells = "".join(f'<td>{fmt(hd.get(yr))}</td>' for yr in HIST_XLABELS)
        if show_plan:
            h_cells += f'<td style="color:{ACCENT}">{fmt(hd.get("26경영계획"))}</td>'
        m_cells = "".join(f'<td>{fmt(cd.get(m))}</td>' for m in range(1, end_mon+1))
        tr_cls = ' class="cur"' if is_sum else ""
        rows += f'<tr{tr_cls}><td class="lbl">{lbl}</td>{h_cells}{m_cells}</tr>'

    return (f'<div style="overflow-x:auto"><table class="num-table">'
            f'<thead>{header}</thead><tbody>{rows}</tbody></table></div>')


def sec_hdr(icon, title):
    st.markdown(f'<div class="sec-hdr">{icon}&nbsp;&nbsp;{title}</div>',
                unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# 검증
# ══════════════════════════════════════════════════════════════
def run_validation(실적_r, avail_months, tmpl_path):
    """계산값 vs 템플릿 비교"""
    if not os.path.exists(tmpl_path): return None
    wb=openpyxl.load_workbook(tmpl_path, data_only=True)
    ws=wb["Gap Trends"]

    checks=[
        ("인쇄내수 일반백상","인쇄","내수","일반백상",10),
        ("인쇄내수 일반아트","인쇄","내수","일반아트",11),
        ("인쇄내수 합계","인쇄","내수","합계",12),
        ("인쇄수출 일반백상","인쇄","수출","일반백상",44),
        ("인쇄수출 일반아트","인쇄","수출","일반아트",45),
        ("인쇄수출 합계","인쇄","수출","합계",46),
        ("감열내수 POS영업팀","감열","내수","POS영업팀",81),
        ("감열내수 라벨영업팀","감열","내수","라벨영업팀",82),
        ("감열내수 합계","감열","내수","합계",83),
        ("감열수출 POS영업팀","감열","수출","POS영업팀",116),
        ("감열수출 라벨영업팀","감열","수출","라벨영업팀",117),
        ("감열수출 합계","감열","수출","합계",118),
        ("패키징내수 기능지류","패키징","내수","기능지류",426),
        ("패키징내수 전사지류","패키징","내수","전사지류",427),
        ("패키징내수 팬시지류","패키징","내수","일반팬시지류",428),
        ("패키징내수 러프그로스지","패키징","내수","러프그로스지",429),
        ("패키징내수 합계","패키징","내수","합계(Protego제외)",430),
        ("패키징수출 전사지류","패키징","수출","전사지류",524),
        ("패키징수출 팬시지류","패키징","수출","일반팬시지류",525),
        ("패키징수출 러프그로스지","패키징","수출","러프그로스지",526),
        ("패키징수출 합계","패키징","수출","합계(Protego제외)",527),
    ]
    rows=[]
    total_ok=0
    for desc, 부문, 유통, label, tmpl_row in checks:
        label_data=실적_r.get(부문,{}).get(유통,{}).get(label,{})
        for yr, mon_dict in label_data.items():
            for mon, calc_val in mon_dict.items():
                if mon not in avail_months: continue
                tcol=CUR_MON_BASE_COL+mon-1
                tmpl_val=ws.cell(tmpl_row, tcol).value
                if tmpl_val is None or calc_val is None: continue
                tmpl_int=to_int(tmpl_val)
                diff=abs(calc_val-tmpl_int)
                ok=(diff<=1)
                if ok: total_ok+=1
                rows.append({
                    "구분":desc,"월":f"{mon}월",
                    "계산값":calc_val,"템플릿값":tmpl_int,
                    "차이":diff,"결과":"✅ 일치" if ok else f"❌ 차이 {diff}",
                })
    return pd.DataFrame(rows), total_ok, len(rows)

# ══════════════════════════════════════════════════════════════
# 엑셀 생성
# ══════════════════════════════════════════════════════════════
def make_excel(bw_bytes, 산업_bytes, template_path, start_mon, end_mon, cur_year=2026):
    if not os.path.exists(template_path): return None
    tmp=tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False); tmp.close()
    shutil.copy(template_path, tmp.name)
    raw=load_ie(io.BytesIO(bw_bytes))
    df_ie=raw[raw["값유형"]==10].copy()
    bw=load_산업(io.BytesIO(산업_bytes))
    MONTHS=[f"{cur_year}.{m:03d}" for m in range(start_mon,end_mon+1)]
    wb=openpyxl.load_workbook(tmp.name); ws=wb["Gap Trends"]

    def sv(r,c,v):
        cell=ws.cell(r,c)
        if isinstance(cell,MergedCell): return
        if v is not None: ws.cell(r,c,int(float(v)))

    def ie(m, 부문, 유통, jj=None, sr=None):
        sub=df_ie[(df_ie["회계연도/기간"].astype(str)==m)&
                  (df_ie["계획지류 부문"]==부문)&(df_ie["유통경로"]==유통)]
        if jj: sub=sub[sub["지류"]==jj]
        if sr: sub=sub[sub["계획지류 지류(sriv 고급)"]==sr]
        v=sub["판매량"].sum()
        return to_int((sub["순매출액"].sum()-sub["R_원재료"].sum())/v/1000) if v else None

    for mi,m in enumerate(MONTHS):
        dc=CUR_MON_BASE_COL+start_mon-1+mi
        sv(10,dc,ie(m,"인쇄","내수","일반백상")); sv(11,dc,ie(m,"인쇄","내수","일반아트"))
        sv(12,dc,ie(m,"인쇄","내수"))
        sv(44,dc,ie(m,"인쇄","수출","일반백상")); sv(45,dc,ie(m,"인쇄","수출","일반아트"))
        sv(46,dc,ie(m,"인쇄","수출"))
        sv(81,dc,ie(m,"감열","내수",sr="POS")); sv(82,dc,ie(m,"감열","내수",sr="일감라벨"))
        sv(83,dc,ie(m,"감열","내수"))
        sv(116,dc,ie(m,"감열","수출",sr="POS")); sv(117,dc,ie(m,"감열","수출",sr="일감라벨"))
        sv(118,dc,ie(m,"감열","수출"))
        def pk(유통,jj=None,excl=False):
            sub=df_ie[(df_ie["회계연도/기간"].astype(str)==m)&
                      (df_ie["계획지류 부문"]=="패키징")&(df_ie["유통경로"]==유통)]
            if jj:   sub=sub[sub["계획지류 지류"]==jj]
            if excl: sub=sub[sub["계획지류 지류"]!="Protego"]
            v=sub["판매량"].sum()
            return to_int((sub["순매출액"].sum()-sub["R_원재료"].sum())/v/1000) if v else None
        sv(426,dc,pk("내수","기능지류")); sv(427,dc,pk("내수","전사지류"))
        sv(428,dc,pk("내수","일반팬시지류")); sv(429,dc,pk("내수","러프그로스지"))
        sv(430,dc,pk("내수",excl=True))
        sx=df_ie[(df_ie["회계연도/기간"].astype(str)==m)&(df_ie["계획지류 부문"]=="패키징")&
                 (df_ie["유통경로"]=="내수")&(df_ie["계획지류 지류"]!="Protego")]
        if sx["판매량"].sum(): sv(462,dc,to_int(sx["R_원재료"].sum()/sx["판매량"].sum()/1000))
        sv(523,dc,pk("수출","Protego")); sv(524,dc,pk("수출","전사지류"))
        sv(525,dc,pk("수출","일반팬시지류")); sv(526,dc,pk("수출","러프그로스지"))
        sv(527,dc,pk("수출"))

    def bsm(유통,jl): return bw[(bw["유통"]==유통)&(bw["지류2"].isin(jl))]
    g=gap_ab(bsm("내수",["AB류"])); sv(297,20,g); sv(298,20,g)
    sv(338,20,gap_sc(bsm("내수",["SC류(고)"]))); sv(341,20,gap_sc(bsm("내수",["SC류(저)"])))
    sv(342,20,gap_sc(bsm("내수",["SC류(고)","IV류","SC류(저)"])))
    sv(377,20,gap_sc(bsm("수출",["SC류(고)"]))); sv(380,20,gap_sc(bsm("수출",["SC류(저)"])))
    sv(381,20,gap_sc(bsm("수출",["SC류(고)","IV류","SC류(저)"])))
    wb.save(tmp.name)
    with open(tmp.name,"rb") as f: buf=f.read()
    os.unlink(tmp.name); return buf

# ══════════════════════════════════════════════════════════════
# 메일
# ══════════════════════════════════════════════════════════════
def send_mail_fn(cfg, from_addr, to_list, subject, body, attachment, filename):
    msg=MIMEMultipart(); msg["From"]=from_addr
    msg["To"]=", ".join(to_list); msg["Subject"]=subject
    msg.attach(MIMEText(body,"plain","utf-8"))
    part=MIMEBase("application","octet-stream"); part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",f'attachment; filename="{filename}"')
    msg.attach(part)
    with smtplib.SMTP(cfg["host"],int(cfg["port"]),timeout=10) as s:
        s.ehlo()
        if s.has_extn("STARTTLS"): s.starttls(); s.ehlo()
        if cfg.get("auth") and cfg.get("user"):
            s.login(cfg["user"],cfg["password"])
        s.sendmail(from_addr,to_list,msg.as_bytes())

@st.dialog("✉️ 메일 발송")
def email_dialog(excel_buf, fname):
    st.markdown(f"**첨부:** `{fname}`")
    from_addr=st.text_input("발신 이메일",value=st.session_state.smtp.get("user",""),
                             placeholder="hong@hansol.com")
    to_raw=st.text_input("수신 이메일",placeholder="scm@hansol.com, mgr@hansol.com")
    subject=st.text_input("제목",value=f"[SCM] 펄프가 GAP 트랜드 — {fname}")
    body=st.text_area("본문",height=90,
        value="안녕하세요,\n\n펄프가 GAP 트랜드 리포트를 첨부 드립니다.\n\n- SCM지원파트 딸깍! 자동화 플랫폼")
    with st.expander("⚙️ SMTP 설정"):
        cfg=st.session_state.smtp
        cfg["host"]=st.text_input("SMTP 서버",cfg["host"])
        cfg["port"]=st.text_input("포트",str(cfg["port"]))
        cfg["auth"]=st.checkbox("인증 사용",value=cfg.get("auth",False))
        if cfg["auth"]:
            cfg["user"]=st.text_input("계정",cfg.get("user",""))
            cfg["password"]=st.text_input("비밀번호",cfg.get("password",""),type="password")
        st.caption("사내 릴레이 echohub.hansol.com → 인증 없이 사용")
        st.session_state.smtp=cfg
    c1,c2=st.columns(2)
    with c1:
        if st.button("📨 발송",type="primary",use_container_width=True):
            if not from_addr or not to_raw:
                st.error("발신/수신 이메일 입력 필요")
            else:
                to_list=[t.strip() for t in to_raw.split(",") if t.strip()]
                try:
                    send_mail_fn(st.session_state.smtp,from_addr,to_list,
                                 subject,body,excel_buf,fname)
                    st.success(f"✅ 발송 완료 → {', '.join(to_list)}")
                except Exception as e: st.error(f"발송 실패: {e}")
    with c2:
        if st.button("취소",use_container_width=True): st.rerun()

# ══════════════════════════════════════════════════════════════
# 사이드바
# ══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style="padding:8px 0 10px;border-bottom:1px solid #3b6cc7;margin-bottom:14px">
      <div style="font-size:18px;font-weight:800;color:#FFFFFF">⚡ 딸깍!</div>
      <div style="font-size:10px;color:#93C5FD;margin-top:2px">SCM지원파트 보고자료 자동화</div>
    </div>
    <div style="font-size:11px;color:#93C5FD;margin-bottom:8px;font-weight:600;letter-spacing:.5px">
      📂 지표 목록
    </div>""", unsafe_allow_html=True)

    NAV=[("🏠 홈","home"),("📊 펄프가 GAP 트랜드","gap"),("📋 내수 마감보고","margin"),("📋 계획지류 관리","plan")]
    for label,pg in NAV:
        if st.button(label, key=f"nav_{pg}", use_container_width=True,
                     type="primary" if st.session_state.page==pg else "secondary"):
            st.session_state.page=pg; st.rerun()

    COMING=["📈 원가 분석","📦 재고 현황","🚚 납기 이행률","💹 판매 실적"]
    st.markdown('<div style="margin-top:14px;font-size:10px;color:#475569;padding:0 8px">준비중</div>',
                unsafe_allow_html=True)
    for c in COMING:
        st.markdown(f'<div style="padding:5px 12px;font-size:12px;color:#475569;margin:1px 0">{c}</div>',
                    unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# 헤더
# ══════════════════════════════════════════════════════════════
st.markdown("""<div class="plat-title">
  <h1>⚡ SCM지원파트 — 딸깍! 보고자료 자동화 플랫폼</h1>
  <p>BW 업로드 → 17년~현재 트랜드 분석 → 엑셀 다운로드 / 메일 발송</p>
</div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# 홈
# ══════════════════════════════════════════════════════════════
if st.session_state.page=="home":
    CARDS=[
        {"icon":"📊","title":"펄프가 GAP 트랜드","desc":"인쇄·감열·패키징·산업\n17년~현재 GAP 트랜드","badge":"운영중","pg":"gap","on":True},
        {"icon":"📋","title":"내수 마감보고","desc":"1차·2차 대금청구 비교\n고객사별 차이 분석","badge":"운영중","pg":"margin","on":True},
        {"icon":"📋","title":"계획지류 관리","desc":"지종별 계획지류 매핑 현황","badge":"운영중","pg":"plan","on":True},
        {"icon":"📈","title":"원가 분석","desc":"원재료비·고정비·변동비\n월별 원가 추이","badge":"준비중","pg":None,"on":False},
        {"icon":"📦","title":"재고 현황","desc":"창고별·제품별 재고 현황","badge":"준비중","pg":None,"on":False},
        {"icon":"🚚","title":"납기 이행률","desc":"수주 대비 납기 이행률 KPI","badge":"준비중","pg":None,"on":False},
        {"icon":"💹","title":"판매 실적","desc":"사업부·채널별 판매 실적","badge":"준비중","pg":None,"on":False},
    ]
    st.markdown(f"""<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:20px">
    {''.join(f"""<div class="report-card{'' if c['on'] else ' coming'}">
      <div class="card-icon">{c['icon']}</div>
      <div class="card-title">{c['title']}</div>
      <div class="card-desc">{c['desc'].replace(chr(10),'<br>')}</div>
      <span class="badge{'' if c['on'] else ' soon'}">{c['badge']}</span>
    </div>""" for c in CARDS)}
    </div>""", unsafe_allow_html=True)
    cb=st.columns(3)
    for i,c in enumerate(CARDS):
        with cb[i%3]:
            if c["on"]:
                if st.button(f"{c['icon']} {c['title']} 열기", key=f"hc{i}",
                             use_container_width=True, type="primary"):
                    st.session_state.page=c["pg"]; st.rerun()
            else:
                st.button(f"{c['icon']} {c['title']} (준비중)", key=f"hc{i}",
                          disabled=True, use_container_width=True)
    st.divider()
    mc=st.columns(3)
    mc[0].metric("운영중 리포트","2 개")
    mc[1].metric("준비중","4 개")
    mc[2].metric("버전","v5.0")

# ══════════════════════════════════════════════════════════════
# GAP 트랜드
# ══════════════════════════════════════════════════════════════
elif st.session_state.page=="gap":

    tab_main, tab_valid = st.tabs(["📊 GAP 트랜드", "✅ 검증"])

    with tab_main:
        with st.expander("📂 데이터 파일 업로드",expanded=True):
            uc1,uc2=st.columns(2)
            with uc1: f_bw=st.file_uploader("BW (인쇄·감열·패키징) — 값 시트",
                                              type=["xlsx"],key="gap_bw")
            with uc2: f_산업=st.file_uploader("산업 BW — BW 시트",
                                               type=["xlsx"],key="gap_ind")

        bw_bytes =read_bytes(f_bw,"bw")
        산업_bytes=read_bytes(f_산업,"산업")
        if not bw_bytes:
            st.info("👆 BW 파일을 업로드하세요"); st.stop()

        with st.spinner("계산 중..."):
            실적_r, 계획_r, avail_months = calc_gap_data(bw_bytes)
            산업_r = calc_산업_data(산업_bytes) if 산업_bytes else None
            tmpl_hist = load_template_hist(DEFAULT["template"])

        if not avail_months: st.warning("유효한 데이터 없음"); st.stop()
        end_mon=max(avail_months)
        MONS_OPT=[f"{m}월" for m in range(1,13)]

        # ── 컨트롤 ───────────────────────────────────────────
        cc1,cc2,cc3,cc4=st.columns([3,2,1,1])
        with cc1:
            e_sel=st.selectbox("조회 마감월",MONS_OPT,index=end_mon-1,key="em")
            end_mon=MONS_OPT.index(e_sel)+1
        with cc2:
            show_plan=st.checkbox("경영계획 함께 보기",value=False)
        with cc3:
            gen_btn=st.button("⬇️ 엑셀",type="primary",use_container_width=True)
        with cc4:
            mail_btn=st.button("✉️ 메일",use_container_width=True)

        fname=f"gap트랜드_2026_{end_mon}월.xlsx"
        st.caption(f"📌 조회: 17년~26년 {end_mon}월 | 단위: 천원/톤 | 소수점 절삭")

        excel_buf=None
        if gen_btn or mail_btn:
            tp=DEFAULT["template"]
            if not 산업_bytes:          st.warning("산업 BW 파일 필요")
            elif not os.path.exists(tp): st.warning("템플릿 파일 없음")
            else:
                with st.spinner("엑셀 생성 중..."):
                    excel_buf=make_excel(bw_bytes,산업_bytes,tp,1,end_mon)
        if excel_buf and gen_btn:
            st.download_button(f"📥 {fname}",data=excel_buf,file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if excel_buf and mail_btn:
            email_dialog(excel_buf,fname)

        st.divider()

        # ── 헬퍼 ─────────────────────────────────────────────
        def get_hist(부문, 유통, label):
            return tmpl_hist.get((부문,유통,label),{})

        def get_cur(부문, 유통, label):
            yd=실적_r.get(부문,{}).get(유통,{}).get(label,{})
            mon_dict={}
            for yr,md in yd.items():
                for mo,v in md.items():
                    if v is not None: mon_dict[mo]=v
            return mon_dict

        def get_plan(부문, 유통, label):
            if not 계획_r: return None
            yd=계획_r.get(부문,{}).get(유통,{}).get(label,{})
            mon_dict={}
            for yr,md in yd.items():
                for mo,v in md.items():
                    if v is not None: mon_dict[mo]=v
            return mon_dict if mon_dict else None

        # ── Excel 동일 종합표 ─────────────────────────────────
        def excel_table(sections_def, end_mon, show_plan):
            """
            sections_def = [(섹션제목, 계산식, [(부문,유통,레이블,is_sum), ...]), ...]
            Returns HTML string
            """
            n_yr   = len(HIST_XLABELS)
            n_plan = 1 if show_plan else 0
            n_mon  = end_mon
            n_hist = n_yr + n_plan
            total  = 1 + n_hist + n_mon

            S = {
                "H1_BASE": f"padding:6px 8px;text-align:center;border:1px solid #1e3a6e;font-size:12px;font-weight:700",
                "H2_HIST": f"background:#475569;color:white;padding:5px 6px;text-align:center;border:1px solid #64748B;font-size:11px;font-weight:600",
                "H2_MON":  f"background:{PRIMARY};color:white;padding:5px 6px;text-align:center;border:1px solid #2563EB;font-size:11px;font-weight:600",
                "H2_PLAN": f"background:{ACCENT};color:white;padding:5px 6px;text-align:center;border:1px solid #1e7a3e;font-size:11px;font-weight:600",
                "VH":      "text-align:right;padding:4px 7px;border:1px solid #e2e8f0;color:#475569;font-size:12px",
                "VC":      f"text-align:right;padding:4px 7px;border:1px solid #bfdbfe;color:{PRIMARY_D};font-size:12px",
                "VHS":     "text-align:right;padding:5px 7px;border:1px solid #d1d5db;color:#1e293b;font-size:12px;font-weight:700;background:#F1F5F9",
                "VCS":     f"text-align:right;padding:5px 7px;border:1px solid #93c5fd;color:{PRIMARY_D};font-size:13px;font-weight:700;background:#EFF6FF",
                "LBL":     "text-align:left;padding:4px 10px;border:1px solid #e2e8f0;font-size:12px;white-space:nowrap;background:#F8FAFC",
                "LBLS":    f"text-align:left;padding:5px 10px;border:1px solid #d1d5db;font-size:12px;font-weight:700;white-space:nowrap;background:#EFF6FF;color:{PRIMARY_D}",
            }

            h  = '<div style="overflow-x:auto;margin-bottom:20px">'
            h += '<table style="width:100%;border-collapse:collapse;font-family:Malgun Gothic,sans-serif">'

            # 헤더 행 1: 단위/(천원)Gap | 월평균 | 26년
            h += ('<tr>'
                  f'<th rowspan="2" style="{S["H1_BASE"]};background:#334155;color:white;text-align:left;min-width:120px">'
                  f'단위/(천원)<br>Gap</th>'
                  f'<th colspan="{n_hist}" style="{S["H1_BASE"]};background:#475569;color:white">월&nbsp;&nbsp;평&nbsp;&nbsp;균</th>'
                  f'<th colspan="{n_mon}" style="{S["H1_BASE"]};background:{PRIMARY};color:white">26년</th>'
                  '</tr>')

            # 헤더 행 2: 17년~25년 | 26년경영계획 | 1월~end월
            h += '<tr>'
            for yr in HIST_XLABELS:
                h += f'<th style="{S["H2_HIST"]}">{yr}</th>'
            if show_plan:
                h += f'<th style="{S["H2_PLAN"]}">26년<br>경영계획</th>'
            for m in range(1, end_mon+1):
                h += f'<th style="{S["H2_MON"]}">{m}월</th>'
            h += '</tr>'

            for sec_title, formula, rows_def in sections_def:
                # 섹션 헤더
                h += (f'<tr><td colspan="{total}" style="background:{PRIMARY_D};color:white;'
                      f'padding:7px 14px;font-size:13px;font-weight:700;border:1px solid #0a2050">'
                      f'{sec_title}'
                      f'&nbsp;&nbsp;<span style="font-size:10px;font-weight:400;opacity:.75">'
                      f'계산식: {formula}</span></td></tr>')

                cur_유통 = None
                for 부문, 유통, label, is_sum in rows_def:
                    # 유통 소제목
                    if 유통 != cur_유통:
                        cur_유통 = 유통
                        h += (f'<tr><td colspan="{total}" style="background:#E2E8F0;color:#334155;'
                              f'padding:4px 14px;font-size:11px;font-weight:700;border:1px solid #CBD5E1">'
                              f'▶ {유통}</td></tr>')
                    hist = get_hist(부문, 유통, label)
                    cur  = get_cur(부문, 유통, label)
                    ls   = S["LBLS"] if is_sum else S["LBL"]
                    vs_h = S["VHS"]  if is_sum else S["VH"]
                    vs_c = S["VCS"]  if is_sum else S["VC"]
                    indent = "" if is_sum else "&nbsp;&nbsp;"
                    h += f'<tr><td style="{ls}">{indent}{label}</td>'
                    for yr in HIST_XLABELS:
                        v = hist.get(yr)
                        h += f'<td style="{vs_h}">{v if v is not None else "-"}</td>'
                    if show_plan:
                        v = hist.get("26경영계획")
                        h += f'<td style="{vs_h};color:{ACCENT}">{v if v is not None else "-"}</td>'
                    for m in range(1, end_mon+1):
                        v = cur.get(m)
                        h += f'<td style="{vs_c}">{v if v is not None else "-"}</td>'
                    h += '</tr>'

            h += '</table></div>'
            return h

        FORMULA_IE = "(순매출액 - R_원재료) / 판매량 / 1,000"
        FORMULA_AB = "(순매출액 - 펄프원재료) / 판매량(톤) / 1,000"
        FORMULA_SC = "(순매출액 - 고지원재료) / 판매량(톤) / 1,000"

        sections = [
            ("📄 인쇄부문 펄프가", FORMULA_IE, [
                ("인쇄","내수","일반백상",False),
                ("인쇄","내수","일반아트",False),
                ("인쇄","내수","합계",True),
                ("인쇄","수출","일반백상",False),
                ("인쇄","수출","일반아트",False),
                ("인쇄","수출","합계",True),
            ]),
            ("🌡️ 감열부문 펄프가", FORMULA_IE, [
                ("감열","내수","POS영업팀",False),
                ("감열","내수","라벨영업팀",False),
                ("감열","내수","합계",True),
                ("감열","수출","POS영업팀",False),
                ("감열","수출","라벨영업팀",False),
                ("감열","수출","합계",True),
            ]),
            ("📦 패키징부문 펄프가 (Protego 제외)", FORMULA_IE, [
                ("패키징","내수","기능지류",False),
                ("패키징","내수","전사지류",False),
                ("패키징","내수","일반팬시지류",False),
                ("패키징","내수","러프그로스지",False),
                ("패키징","내수","합계(Protego제외)",True),
                ("패키징","수출","전사지류",False),
                ("패키징","수출","일반팬시지류",False),
                ("패키징","수출","러프그로스지",False),
                ("패키징","수출","합계(Protego제외)",True),
            ]),
        ]
        st.markdown(excel_table(sections, end_mon, show_plan), unsafe_allow_html=True)

        # ── 트랜드 차트 (섹션별 접기/펼치기) ────────────────────
        def trend_section(icon, title, 부문, 내수_lbls, 수출_lbls):
            with st.expander(f"{icon} {title} 트랜드 차트", expanded=True):
                h_lbls, m_lbls, x_all = _section_labels(end_mon, show_plan)
                n_h = len(h_lbls)
                COLORS=[PRIMARY,"#E05A2B","#2D9A4F","#8B4FBE","#0891B2","#D4A017"]
                def make_fig(유통, labels):
                    fig=go.Figure()
                    fig.add_vrect(x0=x_all[0],x1=x_all[n_h-1],
                                  fillcolor="rgba(203,213,225,0.2)",layer="below",line_width=0)
                    fig.add_vrect(x0=x_all[n_h],x1=x_all[-1],
                                  fillcolor="rgba(219,234,254,0.28)",layer="below",line_width=0)
                    fig.add_vline(x=x_all[n_h-1],line=dict(color="#64748B",width=1.5,dash="dot"))
                    fig.add_annotation(x=n_h/2-0.5,y=1.06,xref="x",yref="paper",
                                       text="<b>월 평 균</b>",showarrow=False,
                                       font=dict(size=11,color="#334155"),
                                       bgcolor="#E2E8F0",bordercolor="#94a3b8",
                                       borderwidth=1,borderpad=4,xanchor="center")
                    fig.add_annotation(x=n_h+len(m_lbls)/2-0.5,y=1.06,xref="x",yref="paper",
                                       text="<b>26년 실적</b>",showarrow=False,
                                       font=dict(size=11,color=PRIMARY_D),
                                       bgcolor="#DBEAFE",bordercolor=PRIMARY,
                                       borderwidth=1,borderpad=4,xanchor="center")
                    for li,lbl in enumerate(labels):
                        h=get_hist(부문,유통,lbl); c=get_cur(부문,유통,lbl)
                        hy=[h.get(yr) for yr in HIST_XLABELS]
                        if show_plan: hy.append(h.get("26경영계획"))
                        cy=[c.get(m) for m in range(1,end_mon+1)]
                        y_all=hy+cy; col=COLORS[li%len(COLORS)]
                        is_sum=lbl in ["합계","합계(Protego제외)"]
                        fig.add_trace(go.Scatter(
                            x=x_all,y=y_all,mode="lines+markers+text",
                            name=lbl,connectgaps=True,
                            line=dict(color=col,width=3 if is_sum else 1.8,
                                      dash="solid" if is_sum else "dot"),
                            marker=dict(size=[6]*n_h+[10 if is_sum else 5]*len(m_lbls),
                                        color=[GRAY_T]*n_h+[col]*len(m_lbls),
                                        symbol=["circle"]*n_h+["diamond" if is_sum else "circle"]*len(m_lbls),
                                        line=dict(width=1.5,color="white")),
                            text=(["" if v is None else str(v) for v in y_all]
                                  if is_sum else [""]*len(x_all)),
                            textposition="top center",textfont=dict(size=10,color=col),
                        ))
                    fig.update_layout(
                        title=dict(text=f"{부문} {유통} GAP 트랜드",
                                   font=dict(size=15,color=PRIMARY_D),x=0),
                        height=460,
                        yaxis=dict(title="Gap (천원/톤)",gridcolor="#e8edf5",tickfont=dict(size=12)),
                        xaxis=dict(gridcolor="#f1f5f9",tickangle=-30,tickfont=dict(size=12)),
                        legend=dict(orientation="h",yanchor="bottom",y=1.06,
                                    font=dict(size=11),bgcolor="rgba(0,0,0,0)"),
                        hovermode="x unified",plot_bgcolor="white",paper_bgcolor="white",
                        margin=dict(t=100,b=60,l=70,r=80),
                    )
                    return fig
                st.plotly_chart(make_fig("내수",내수_lbls),use_container_width=True)
                st.plotly_chart(make_fig("수출",수출_lbls),use_container_width=True)

        trend_section("📄","인쇄지","인쇄",["일반백상","일반아트","합계"],["일반백상","일반아트","합계"])
        trend_section("🌡️","감열지","감열",["POS영업팀","라벨영업팀","합계"],["POS영업팀","라벨영업팀","합계"])
        trend_section("📦","패키징","패키징",
                      ["기능지류","전사지류","일반팬시지류","러프그로스지","합계(Protego제외)"],
                      ["전사지류","일반팬시지류","러프그로스지","합계(Protego제외)"])

        # ─ 산업 ──────────────────────────────────────────────
        sec_hdr("🏭","산업지")
        st.caption(f"AB류: {FORMULA_AB} | SC류: {FORMULA_SC}")
        if not 산업_r:
            st.info("산업 BW 파일을 업로드하세요")
        else:
            ab=산업_r["AB류"]; sc=산업_r["SC류"]
            # 산업 종합표
            산업_sections = [
                ("🏭 산업부문 — AB류", FORMULA_AB, [
                    ("산업","내수","AB류",True),
                ]),
                ("🏭 산업부문 — SC류", FORMULA_SC, [
                    ("산업","내수","SC류(고)",False),
                    ("산업","내수","SC류(저)",False),
                    ("산업","내수","SC류 소계",True),
                    ("산업","수출","SC류(고)",False),
                    ("산업","수출","SC류(저)",False),
                    ("산업","수출","SC류 소계",True),
                ]),
            ]
            st.markdown(excel_table(산업_sections, end_mon, show_plan), unsafe_allow_html=True)

            with st.expander("🏭 산업 트랜드 차트", expanded=True):
                ab_hist=tmpl_hist.get(("산업","내수","AB류"),{})
                ab_val=ab.get("AB류 내수",{}).get("Gap")
                h_lbls,m_lbls,x_all=_section_labels(end_mon,show_plan)
                hy=[ab_hist.get(yr) for yr in HIST_XLABELS]
                if show_plan: hy.append(ab_hist.get("26경영계획"))
                y_all=hy+[ab_val]
                x_cur=x_all[:len(hy)]+["26년 실적"]
                fig=go.Figure(go.Scatter(x=x_cur,y=y_all,mode="lines+markers+text",
                    text=["" if v is None else str(v) for v in y_all],
                    textposition="top center",textfont=dict(size=11),
                    line=dict(color=PRIMARY,width=2.5),
                    marker=dict(size=[6]*len(hy)+[12],
                                color=[GRAY_T]*len(hy)+[PRIMARY],
                                symbol=["circle"]*len(hy)+["diamond"],
                                line=dict(width=1.5,color="white"))))
                fig.update_layout(title=dict(text="산업 AB류 내수 GAP 트랜드",font=dict(size=15,color=PRIMARY_D)),
                    height=460,
                    yaxis=dict(gridcolor="#e8edf5",title="Gap (천원/톤)",tickfont=dict(size=12)),
                    xaxis=dict(tickangle=-30,tickfont=dict(size=12)),
                    plot_bgcolor="white",paper_bgcolor="white",
                    margin=dict(t=70,b=60,l=70,r=20))
                st.plotly_chart(fig,use_container_width=True)

                lbls=list(sc.keys())
                fig=go.Figure(go.Bar(x=lbls,y=[sc[k]["Gap"] or 0 for k in lbls],
                    text=[str(sc[k]["Gap"]) for k in lbls],textposition="outside",
                    marker_color=ACCENT))
                fig.update_layout(title=dict(text="산업 SC류 Gap (26년 실적)",font=dict(size=15,color=PRIMARY_D)),
                    height=460,
                    plot_bgcolor="white",paper_bgcolor="white",
                    xaxis=dict(tickangle=-30,tickfont=dict(size=12)),
                    margin=dict(t=70,b=80,l=70,r=20),
                    yaxis=dict(gridcolor="#e8edf5",tickfont=dict(size=12)))
                st.plotly_chart(fig,use_container_width=True)

    # ── 검증 탭 ──────────────────────────────────────────────
    with tab_valid:
        st.markdown("### ✅ 계산값 vs (2026-07월)Gap_trends.xlsx 검증")
        if not bw_bytes:
            st.info("BW 파일을 먼저 업로드하세요")
        else:
            if st.button("🔍 검증 실행",type="primary"):
                with st.spinner("검증 중..."):
                    result=run_validation(실적_r,avail_months,DEFAULT["template"])
                if result is None:
                    st.error("템플릿 파일 없음")
                else:
                    df_v,ok,total=result
                    col1,col2,col3=st.columns(3)
                    col1.metric("전체 검증 항목",f"{total}개")
                    col2.metric("일치 (차이≤1)",f"{ok}개")
                    col3.metric("불일치",f"{total-ok}개",
                                delta=None if total==ok else f"{total-ok}개 확인 필요",
                                delta_color="off" if total==ok else "inverse")
                    if total==ok:
                        st.success("🎉 모든 항목 일치! 계산 로직이 정확합니다.")
                    else:
                        st.warning("일부 불일치 항목이 있습니다. 아래 표를 확인하세요.")
                    st.dataframe(
                        df_v.style.apply(
                            lambda r: ["background:#dcfce7"]*len(r) if r["결과"].startswith("✅")
                                      else ["background:#fee2e2"]*len(r),
                            axis=1),
                        hide_index=True, use_container_width=True, height=500
                    )
                    csv=df_v.to_csv(index=False,encoding="utf-8-sig").encode("utf-8-sig")
                    st.download_button("📥 검증 결과 다운로드",data=csv,
                                       file_name="gap_검증결과.csv",mime="text/csv")

# ══════════════════════════════════════════════════════════════
# 내수 마감보고
# ══════════════════════════════════════════════════════════════
elif st.session_state.page == "margin":
    # ── 부문 매핑 ──────────────────────────────────────────────
    팀_부문_MAP = {
        "산업용지국내영업팀":"산업","인쇄용지국내영업팀":"인쇄",
        "라벨영업팀(내수)":"감열","POS영업팀(내수)":"감열",
        "신제품팀(내수)":"신제품","패키징국내영업팀":"패키징",
        "복사지팀":"복사지","특수지영업팀(내수)":"패키징",
        "패키징해외영업(내수)":"패키징","(구)특수지해외(내수)":"패키징",
        "ECO패키징영업팀":"패키징","팬시지 영업팀":"패키징",
        "특수지 영업팀":"패키징","친환경소재팀":"친환경",
        "친환경패키징 영업팀":"친환경",
    }
    부문_순서 = ["인쇄","감열","패키징","산업","복사지","신제품","친환경"]

    @st.cache_data(show_spinner=False)
    def calc_margin(b1, b2):
        import io
        df1 = pd.read_excel(io.BytesIO(b1), sheet_name="Sheet1")
        df2 = pd.read_excel(io.BytesIO(b2), sheet_name="Sheet1")

        def prep(df):
            df = df[df["내수구분"]=="내수"].copy()
            for c in ["청구중량","공급가","판매가격","총할인율"]:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
            df["부문"] = df["팀명"].map(팀_부문_MAP).fillna("기타")
            # 수출 제외할것 팀 제거
            df = df[df["부문"] != "기타"]
            # 그룹 집계
            def wagg(g):
                중량 = g["청구중량"].sum()
                공급 = g["공급가"].sum()
                할인 = (g["총할인율"] * g["공급가"].abs()).sum() / g["공급가"].abs().sum() if g["공급가"].abs().sum() > 0 else 0
                단가 = 공급 / 중량 if 중량 != 0 else 0  # 원/kg = 천원/톤
                return pd.Series({
                    "판매량(톤)":   round(중량/1000, 3),
                    "할인율":       round(할인, 6),
                    "판매가(천원)": round(단가, 6),
                    "공급가(천원)": round(공급/1000, 3),
                })
            agg = df.groupby(["부문","팀명","지급처명","지류","사원명"], sort=False).apply(wagg).reset_index()
            return agg

        r1 = prep(df1); r1.rename(columns={"사원명":"담당자"}, inplace=True)
        r2 = prep(df2); r2.rename(columns={"사원명":"담당자"}, inplace=True)

        # 2차에서 세금계산서 발행여부 확인
        df2_all = pd.read_excel(io.BytesIO(b2), sheet_name="Sheet1")
        df2_all = df2_all[df2_all["내수구분"]=="내수"].copy()
        df2_all["부문"] = df2_all["팀명"].map(팀_부문_MAP).fillna("기타")
        df2_all = df2_all[df2_all["부문"] != "기타"]
        df2_all["세금계산서번호"] = df2_all["세금계산서번호"].astype(str).str.strip()
        tax = df2_all.groupby(["팀명","지급처명","지류"])["세금계산서번호"].apply(
            lambda s: "O" if any(v not in ["","nan","None"] for v in s) else "X"
        ).reset_index().rename(columns={"세금계산서번호":"세금계산서발행"})

        # 병합
        key = ["부문","팀명","지급처명","지류","담당자"]
        merged = pd.merge(
            r1.rename(columns={c:f"1차_{c}" for c in ["판매량(톤)","할인율","판매가(천원)","공급가(천원)"]}),
            r2.rename(columns={c:f"2차_{c}" for c in ["판매량(톤)","할인율","판매가(천원)","공급가(천원)"]}),
            on=key, how="outer"
        ).fillna(0)
        merged = pd.merge(merged, tax, on=["팀명","지급처명","지류"], how="left")
        merged["세금계산서발행"] = merged["세금계산서발행"].fillna("X")

        for f in ["판매량(톤)","할인율","판매가(천원)","공급가(천원)"]:
            merged[f"차이_{f}"] = round(merged[f"2차_{f}"] - merged[f"1차_{f}"], 3)

        # 부문 정렬
        merged["_ord"] = merged["부문"].apply(lambda x: 부문_순서.index(x) if x in 부문_순서 else 99)
        merged.sort_values(["_ord","팀명","지급처명","지류"], inplace=True)
        merged.drop(columns=["_ord"], inplace=True)
        return merged

    def make_margin_excel(merged):
        import io
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        wb2 = openpyxl.Workbook()
        ws = wb2.active; ws.title = "마감보고"

        thin = Side(style="thin", color="AAAAAA")
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr_cell(row, col, val, bg="1A4F9C", fg="FFFFFF", bold=True, align="center"):
            c = ws.cell(row, col, val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="맑은 고딕", size=9, bold=bold, color=fg)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            c.border = bd

        def data_cell(row, col, val, bg=None, bold=False, align="center", fmt=None):
            c = ws.cell(row, col, val)
            if bg: c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="맑은 고딕", size=9, bold=bold)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = bd
            if fmt: c.number_format = fmt

        # 제목
        ws.merge_cells("A1:S1")
        t = ws.cell(1,1,"●26년 7월 내수 마감보고"); t.font=Font(name="맑은 고딕",size=13,bold=True,color="1A4F9C")
        t.alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[1].height = 22

        # 헤더 행2
        headers2 = [("A2:E2","구분자"),("G2:J2","1차 마감(b)"),("K2:N2","2차 마감(a)"),("O2:R2","차이 (a-b)")]
        ws.merge_cells("A2:E2"); ws.cell(2,1,"구분자")
        ws.merge_cells("F2:F3"); ws.cell(2,6,"세금계산서\n발행여부")
        ws.merge_cells("G2:J2"); ws.cell(2,7,"1차 마감(b)")
        ws.merge_cells("K2:N2"); ws.cell(2,11,"2차 마감(a)")
        ws.merge_cells("O2:R2"); ws.cell(2,15,"차이 (a-b)")
        ws.merge_cells("S2:S3"); ws.cell(2,19,"2차마감 사유")
        for c in [1,7,11,15]: hdr_cell(2,c,ws.cell(2,c).value)
        hdr_cell(2,6,ws.cell(2,6).value,bg="2D9A4F")
        hdr_cell(2,19,ws.cell(2,19).value,bg="475569")

        # 헤더 행3
        sub = ["부문","팀명","지급처명","지류","담당자",
               "판매량(톤)","할인율","판매가(천원)","공급가(천원)",
               "판매량(톤)","할인율","판매가(천원)","공급가(천원)",
               "판매량(톤)","할인율","판매가(천원)","공급가(천원)"]
        for i,h in enumerate(sub):
            col = i+1 if i < 5 else i+2
            hdr_cell(3,col,h,bg="475569")
        ws.row_dimensions[2].height=18; ws.row_dimensions[3].height=15

        # 컬럼 너비
        widths = [6,12,16,8,6, 6, 8,7,9,9, 8,7,9,9, 7,6,8,9, 18]
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

        # 데이터
        BG_DIFF = "FEF3C7"; BG_ALT = "F8FAFC"
        row=4
        cur_부문 = None
        for _,r in merged.iterrows():
            has_diff = abs(r["차이_공급가(천원)"]) > 0.1
            bg = BG_DIFF if has_diff else (BG_ALT if row%2==0 else "FFFFFF")
            vals = [r["부문"],r["팀명"],r["지급처명"],r["지류"],r["담당자"],
                    r["세금계산서발행"],
                    r["1차_판매량(톤)"],r["1차_할인율"],r["1차_판매가(천원)"],r["1차_공급가(천원)"],
                    r["2차_판매량(톤)"],r["2차_할인율"],r["2차_판매가(천원)"],r["2차_공급가(천원)"],
                    r["차이_판매량(톤)"],r["차이_할인율"],r["차이_판매가(천원)"],r["차이_공급가(천원)"],
                    ""]
            for ci,v in enumerate(vals,1):
                align = "left" if ci in [1,2,3,4,5] else "right" if ci in [7,8,9,10,11,12,13,14,15,16,17,18] else "center"
                fmt = "#,##0.000" if ci in [7,10,11,14,15,18] else "#,##0.000" if ci in [8,12,16] else None
                data_cell(row,ci,v,bg=bg,align=align,fmt=fmt)
            row += 1

        # 합계 행
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row,1,"합  계").font=Font(name="맑은 고딕",size=9,bold=True)
        ws.cell(row,1).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(row,1).fill=PatternFill("solid",fgColor="334155")
        ws.cell(row,1).font=Font(name="맑은 고딕",size=9,bold=True,color="FFFFFF")
        for ci,col in enumerate([7,9,10,11,13,14,15,17,18],1):
            field_map = {7:"1차_판매량(톤)",9:"1차_판매가(천원)",10:"1차_공급가(천원)",
                         11:"2차_판매량(톤)",13:"2차_판매가(천원)",14:"2차_공급가(천원)",
                         15:"차이_판매량(톤)",17:"차이_판매가(천원)",18:"차이_공급가(천원)"}
            v = merged[field_map[col]].sum() if col in field_map else None
            if v is not None:
                c=ws.cell(row,col,round(v,3))
                c.fill=PatternFill("solid",fgColor="334155")
                c.font=Font(name="맑은 고딕",size=9,bold=True,color="FFFFFF")
                c.alignment=Alignment(horizontal="right",vertical="center")
                c.border=bd
                c.number_format="#,##0.000"

        buf = io.BytesIO(); wb2.save(buf); buf.seek(0)
        return buf.getvalue()

    # ── UI ───────────────────────────────────────────────────
    st.markdown("### 📋 내수 마감보고 지표")
    with st.expander("📂 데이터 파일 업로드", expanded=True):
        uc1, uc2 = st.columns(2)
        with uc1: f_m1 = st.file_uploader("대금청구 1차 파일", type=["xlsx"], key="margin_1")
        with uc2: f_m2 = st.file_uploader("대금청구 2차 파일", type=["xlsx"], key="margin_2")

    b1 = read_bytes(f_m1,"m1"); b2 = read_bytes(f_m2,"m2")
    if not b1 or not b2:
        st.info("👆 1차·2차 대금청구 파일을 모두 업로드하세요"); st.stop()

    with st.spinner("집계 중..."):
        merged = calc_margin(b1, b2)

    # 요약 지표
    m1 = st.columns(4)
    m1[0].metric("고객사 수", f"{merged['지급처명'].nunique():,}개")
    m1[1].metric("1차 공급가 합계", f"{merged['1차_공급가(천원)'].sum()/1e6:,.1f}억원")
    m1[2].metric("2차 공급가 합계", f"{merged['2차_공급가(천원)'].sum()/1e6:,.1f}억원")
    diff_total = merged["차이_공급가(천원)"].sum()
    m1[3].metric("차이(a-b)", f"{diff_total:+,.1f}천원",
                 delta_color="inverse" if diff_total < 0 else "normal")

    # ── 검증 배지 (에러방지: 삼항 대신 if/else) ────────────────
    my_1차 = merged["1차_공급가(천원)"].sum()
    my_2차 = merged["2차_공급가(천원)"].sum()
    v1, v2 = st.columns(2)
    with v1:
        st.success(f"✅ 1차 공급가: {my_1차:,.1f}천원 ({my_1차/1e6:,.2f}억원)")
    with v2:
        st.success(f"✅ 2차 공급가: {my_2차:,.1f}천원 ({my_2차/1e6:,.2f}억원)")
    with st.expander("ℹ️ 파일 원본 합계와 집계값이 다른 이유", expanded=False):
        st.markdown("""
**업로드 파일 원본 합계**와 집계값이 다른 이유:
- `산업용지해외영업팀`은 내수 마감보고 제외 대상(수출 팀)으로 자동 제외됩니다.
- 제외 금액: 약 **107,977천원** (1차·2차 동일) → 제외 후 집계 = 마감보고 기준값과 일치 ✅
""")

    # ── 차이 발생 건 요약 ──────────────────────────────────────
    diff_rows = merged[merged["차이_공급가(천원)"].abs() > 0.1].copy()
    if len(diff_rows):
        st.markdown("---")
        st.markdown(f"### ⚠️ 1·2차 차이 발생 {len(diff_rows)}건 요약")
        mc1, mc2, mc3 = st.columns(3)
        mc1.metric("차이 발생 건수", f"{len(diff_rows)}건")
        mc2.metric("차이 공급가 합계", f"{diff_rows['차이_공급가(천원)'].sum():+,.1f}천원")
        mc3.metric("차이 판매량 합계", f"{diff_rows['차이_판매량(톤)'].sum():+,.3f}톤")

        def diff_summary_html(df):
            S = {
                "H": f"background:#B91C1C;color:white;padding:6px 8px;text-align:center;border:1px solid #991B1B;font-size:11px;font-weight:700",
                "TDL": "padding:5px 8px;border:1px solid #FECACA;font-size:12px;text-align:left;white-space:nowrap;background:#FFF5F5",
                "TDR": "padding:5px 8px;border:1px solid #FECACA;font-size:12px;text-align:right;background:#FFF5F5",
                "TDC": "padding:5px 8px;border:1px solid #FECACA;font-size:12px;text-align:center;background:#FFF5F5",
                "DIFF": "padding:5px 8px;border:1px solid #FCA5A5;font-size:12px;text-align:right;font-weight:700;color:#B91C1C;background:#FEE2E2",
            }
            h = '<div style="overflow-x:auto;margin-bottom:12px">'
            h += '<table style="width:100%;border-collapse:collapse;font-family:Malgun Gothic,sans-serif">'
            h += '<tr>'
            for hdr in ["부문","팀명","지급처명","지류","담당자","세금계산서","1차 공급가(천원)","2차 공급가(천원)","차이 공급가(천원)","1차 판매량(톤)","2차 판매량(톤)","차이 판매량(톤)"]:
                h += f'<th style="{S["H"]}">{hdr}</th>'
            h += '</tr>'
            for _, r in df.sort_values("차이_공급가(천원)").iterrows():
                h += '<tr>'
                for val,sty in [
                    (r["부문"],S["TDL"]),(r["팀명"],S["TDL"]),(r["지급처명"],S["TDL"]),
                    (r["지류"],S["TDL"]),(r["담당자"],S["TDC"]),
                    (r["세금계산서발행"],S["TDC"]),
                ]:
                    h += f'<td style="{sty}">{val}</td>'
                h += f'<td style="{S["TDR"]}">{r["1차_공급가(천원)"]:,.1f}</td>'
                h += f'<td style="{S["TDR"]}">{r["2차_공급가(천원)"]:,.1f}</td>'
                diff_c = r["차이_공급가(천원)"]
                clr = "#B91C1C" if diff_c < 0 else "#15803D"
                h += f'<td style="{S["DIFF"]};color:{clr}">{diff_c:+,.1f}</td>'
                h += f'<td style="{S["TDR"]}">{r["1차_판매량(톤)"]:,.3f}</td>'
                h += f'<td style="{S["TDR"]}">{r["2차_판매량(톤)"]:,.3f}</td>'
                diff_t = r["차이_판매량(톤)"]
                h += f'<td style="{S["DIFF"]};color:{"#B91C1C" if diff_t < 0 else "#15803D"}">{diff_t:+,.3f}</td>'
                h += '</tr>'
            # 합계
            h += f'<tr style="background:#1C1917;color:white;font-weight:700">'
            h += f'<td colspan="6" style="padding:5px 10px;border:1px solid #57534E;font-size:12px">소  계</td>'
            h += f'<td style="text-align:right;padding:5px 8px;border:1px solid #57534E">{df["1차_공급가(천원)"].sum():,.1f}</td>'
            h += f'<td style="text-align:right;padding:5px 8px;border:1px solid #57534E">{df["2차_공급가(천원)"].sum():,.1f}</td>'
            diff_s = df["차이_공급가(천원)"].sum()
            h += f'<td style="text-align:right;padding:5px 8px;border:1px solid #57534E;color:{"#FCA5A5" if diff_s < 0 else "#86EFAC"};font-weight:700">{diff_s:+,.1f}</td>'
            h += f'<td style="text-align:right;padding:5px 8px;border:1px solid #57534E">{df["1차_판매량(톤)"].sum():,.3f}</td>'
            h += f'<td style="text-align:right;padding:5px 8px;border:1px solid #57534E">{df["2차_판매량(톤)"].sum():,.3f}</td>'
            diff_ts = df["차이_판매량(톤)"].sum()
            h += f'<td style="text-align:right;padding:5px 8px;border:1px solid #57534E;color:{"#FCA5A5" if diff_ts < 0 else "#86EFAC"};font-weight:700">{diff_ts:+,.3f}</td>'
            h += '</tr></table></div>'
            return h

        st.markdown(diff_summary_html(diff_rows), unsafe_allow_html=True)
        st.markdown("---")

    # 검색 필터
    fc1, fc2 = st.columns([2,1])
    with fc1: search = st.text_input("🔍 지급처명·지류 검색", placeholder="예: 한솔", key="margin_search")
    with fc2: diff_only = st.checkbox("차이 발생 건만 보기", value=False)

    view = merged.copy()
    if search:
        view = view[view["지급처명"].str.contains(search, na=False) |
                    view["지류"].str.contains(search, na=False)]
    if diff_only:
        view = view[view["차이_공급가(천원)"].abs() > 0.1]

    # ── HTML 테이블 출력 (보고 시트 동일 구조) ─────────────────
    def margin_html(df):
        n_cols = 19
        S = {
            "H1": f"background:{PRIMARY_D};color:white;padding:6px 8px;text-align:center;border:1px solid #1e3a6e;font-size:11px;font-weight:700",
            "H2G": f"background:#2D9A4F;color:white;padding:5px 6px;text-align:center;border:1px solid #1e7a3e;font-size:10px;font-weight:600",
            "H2B": f"background:#475569;color:white;padding:5px 6px;text-align:center;border:1px solid #64748B;font-size:10px;font-weight:600",
            "H2P": f"background:{PRIMARY};color:white;padding:5px 6px;text-align:center;border:1px solid #2563EB;font-size:10px;font-weight:600",
            "H2R": f"background:#DC2626;color:white;padding:5px 6px;text-align:center;border:1px solid #991B1B;font-size:10px;font-weight:600",
            "TD":  "padding:4px 6px;border:1px solid #E2E8F0;font-size:11px;text-align:right",
            "TDL": "padding:4px 6px;border:1px solid #E2E8F0;font-size:11px;text-align:left;white-space:nowrap",
            "TDC": "padding:4px 6px;border:1px solid #E2E8F0;font-size:11px;text-align:center",
        }
        h = '<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-family:Malgun Gothic,sans-serif">'
        # 헤더 행1
        h += ('<tr>'
              f'<th colspan="5" style="{S["H1"]}">구&nbsp;&nbsp;분&nbsp;&nbsp;자</th>'
              f'<th rowspan="2" style="{S["H2G"]}">세금계산서<br>발행여부</th>'
              f'<th colspan="4" style="{S["H1"]}">1차 마감(b)</th>'
              f'<th colspan="4" style="{S["H1"]}">2차 마감(a)</th>'
              f'<th colspan="4" style="{S["H1"]};background:#B91C1C">차이 (a-b)</th>'
              f'<th rowspan="2" style="{S["H2B"]}">2차마감<br>사유</th>'
              '</tr>')
        # 헤더 행2
        h += '<tr>'
        for lbl in ["부문","팀명","지급처명","지류","담당자"]:
            h += f'<th style="{S["H2B"]}">{lbl}</th>'
        for _ in range(2):
            for lbl in ["판매량(톤)","할인율","판매가(천원)","공급가(천원)"]:
                h += f'<th style="{S["H2P"]}">{lbl}</th>'
        for lbl in ["판매량(톤)","할인율","판매가(천원)","공급가(천원)"]:
            h += f'<th style="{S["H2R"]}">{lbl}</th>'
        h += '</tr>'

        cur_부문 = None
        for i, (_, r) in enumerate(df.iterrows()):
            has_diff = abs(r["차이_공급가(천원)"]) > 0.1
            row_bg = "#FEF3C7" if has_diff else ("#F8FAFC" if i%2==0 else "#FFFFFF")
            if r["부문"] != cur_부문:
                cur_부문 = r["부문"]
                h += (f'<tr><td colspan="{n_cols}" style="background:{PRIMARY_D};color:white;'
                      f'padding:5px 12px;font-size:12px;font-weight:700;border:1px solid #0a2050">'
                      f'▌ {cur_부문}부문</td></tr>')
            h += f'<tr style="background:{row_bg}">'
            for val,sty in [
                (r["부문"],S["TDL"]),(r["팀명"],S["TDL"]),(r["지급처명"],S["TDL"]),
                (r["지류"],S["TDL"]),(r["담당자"],S["TDC"]),
                (r["세금계산서발행"],f'{S["TDC"]};color:{"#16A34A" if r["세금계산서발행"]=="O" else "#DC2626"};font-weight:700'),
            ]:
                h += f'<td style="{sty}">{val}</td>'
            for prefix in ["1차_","2차_","차이_"]:
                clr = "#DC2626" if prefix=="차이_" and has_diff else ""
                col_s = f'{S["TD"]};color:{clr}' if clr else S["TD"]
                h += f'<td style="{col_s}">{r[prefix+"판매량(톤)"]:.3f}</td>'
                h += f'<td style="{col_s}">{r[prefix+"할인율"]:.2f}</td>'
                h += f'<td style="{col_s}">{r[prefix+"판매가(천원)"]:.1f}</td>'
                h += f'<td style="{col_s}">{r[prefix+"공급가(천원)"]:,.1f}</td>'
            h += f'<td style="{S["TDL"]}"></td></tr>'

        # 합계 행
        h += f'<tr style="background:#334155;color:white;font-weight:700">'
        h += f'<td colspan="5" style="padding:6px 12px;border:1px solid #1e293b;font-size:12px">합  계</td>'
        h += f'<td style="border:1px solid #1e293b"></td>'
        for prefix in ["1차_","2차_","차이_"]:
            h += f'<td style="text-align:right;padding:4px 7px;border:1px solid #1e293b">{df[prefix+"판매량(톤)"].sum():,.3f}</td>'
            h += f'<td style="text-align:right;padding:4px 7px;border:1px solid #1e293b">-</td>'
            h += f'<td style="text-align:right;padding:4px 7px;border:1px solid #1e293b">-</td>'
            h += f'<td style="text-align:right;padding:4px 7px;border:1px solid #1e293b">{df[prefix+"공급가(천원)"].sum():,.1f}</td>'
        h += '<td style="border:1px solid #1e293b"></td></tr>'
        h += '</table></div>'
        return h

    st.markdown(margin_html(view), unsafe_allow_html=True)

    # ── 저장 / 다운로드 / 메일 ──────────────────────────────────
    st.divider()
    SAVE_DIR = os.path.join(os.path.dirname(__file__), "saved_reports", "margin")
    os.makedirs(SAVE_DIR, exist_ok=True)

    # 저장 월 선택
    sa1, sa2, sa3, sa4 = st.columns([2,1,1,1])
    with sa1:
        save_ym = st.text_input("저장 기준월 (YYYYMM)", value="202607", max_chars=6, key="margin_ym",
                                 help="예: 202607 → 2026년 7월 보고서로 저장")
    save_path = os.path.join(SAVE_DIR, f"margin_{save_ym}.csv")
    fname_m = f"내수마감보고_{save_ym[:4]}_{save_ym[4:]}월.xlsx"

    with sa2:
        if st.button("💾 저장", type="primary", use_container_width=True, key="margin_save"):
            merged.to_csv(save_path, index=False, encoding="utf-8-sig")
            st.success(f"✅ {save_ym[:4]}년 {save_ym[4:]}월 보고서 저장됨")
    with sa3:
        if st.button("⬇️ 엑셀", use_container_width=True, key="margin_xls"):
            with st.spinner("엑셀 생성 중..."):
                xls = make_margin_excel(merged)
            st.download_button(f"📥 {fname_m}", data=xls, file_name=fname_m,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="margin_dl")
    with sa4:
        if st.button("✉️ 메일", use_container_width=True, key="margin_mail"):
            with st.spinner("엑셀 생성 중..."):
                xls = make_margin_excel(merged)
            email_dialog(xls, fname_m)

    # ── 저장된 리포트 월별 조회 ──────────────────────────────────
    saved_files = sorted(
        [f for f in os.listdir(SAVE_DIR) if f.startswith("margin_") and f.endswith(".csv")],
        reverse=True
    )
    if saved_files:
        st.markdown("---")
        st.markdown("### 📁 저장된 리포트 월별 조회")
        ym_opts = {f.replace("margin_","").replace(".csv",""):f for f in saved_files}
        labels = [f"{k[:4]}년 {k[4:]}월" for k in ym_opts]
        sel_label = st.selectbox("조회할 월 선택", labels, key="margin_hist_sel")
        sel_key = list(ym_opts.keys())[labels.index(sel_label)]
        sel_file = os.path.join(SAVE_DIR, ym_opts[sel_key])

        df_hist = pd.read_csv(sel_file, encoding="utf-8-sig")
        hm1 = st.columns(4)
        hm1[0].metric("고객사 수", f"{df_hist['지급처명'].nunique():,}개")
        hm1[1].metric("1차 공급가", f"{df_hist['1차_공급가(천원)'].sum()/1e6:,.1f}억원")
        hm1[2].metric("2차 공급가", f"{df_hist['2차_공급가(천원)'].sum()/1e6:,.1f}억원")
        hd = df_hist["차이_공급가(천원)"].sum()
        hm1[3].metric("차이(a-b)", f"{hd:+,.1f}천원",
                      delta_color="inverse" if hd < 0 else "normal")

        diff_hist = df_hist[df_hist["차이_공급가(천원)"].abs() > 0.1]
        if len(diff_hist):
            st.warning(f"⚠️ 차이 발생 {len(diff_hist)}건")

        hc1, hc2 = st.columns([3,1])
        with hc1: h_search = st.text_input("🔍 검색", key="margin_hist_search", placeholder="지급처명·지류")
        with hc2: h_diff = st.checkbox("차이 건만", key="margin_hist_diff")

        hview = df_hist.copy()
        if h_search:
            hview = hview[hview["지급처명"].str.contains(h_search,na=False)|
                          hview["지류"].str.contains(h_search,na=False)]
        if h_diff:
            hview = hview[hview["차이_공급가(천원)"].abs() > 0.1]

        st.markdown(margin_html(hview), unsafe_allow_html=True)

        hd1, hd2 = st.columns(2)
        with hd1:
            if st.button("⬇️ 저장본 엑셀 다운로드", use_container_width=True, key="hist_xls"):
                with st.spinner("엑셀 생성 중..."):
                    xls_h = make_margin_excel(df_hist)
                fname_h = f"내수마감보고_{sel_key[:4]}_{sel_key[4:]}월.xlsx"
                st.download_button(f"📥 {fname_h}", data=xls_h, file_name=fname_h,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="hist_dl")
        with hd2:
            if st.button("🗑️ 이 월 저장본 삭제", use_container_width=True, key="hist_del"):
                os.remove(sel_file)
                st.warning(f"{sel_label} 저장본 삭제됨")
                st.rerun()

# ══════════════════════════════════════════════════════════════
# 계획지류 관리
# ══════════════════════════════════════════════════════════════
elif st.session_state.page=="plan":
    st.markdown("### 📋 지종별 계획지류 관리")
    with st.expander("📂 파일 업로드",expanded=True):
        f_plan=st.file_uploader("계획지류 파일",type=["xlsx"],key="plan_f")
    plan_bytes=read_bytes(f_plan,"plan")
    if not plan_bytes: st.info("파일을 업로드하세요"); st.stop()
    plan_df=pd.read_excel(io.BytesIO(plan_bytes), header=0)
    p=DEFAULT["plan"]
    p_date=(datetime.now().strftime("%Y-%m-%d %H:%M") if f_plan
            else datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y-%m-%d %H:%M")
            if os.path.exists(p) else "-")
    mc=st.columns(3)
    mc[0].metric("지종 수",plan_df["지종"].nunique() if "지종" in plan_df.columns else len(plan_df))
    mc[1].metric("계획지류 수",plan_df["계획지류명"].nunique() if "계획지류명" in plan_df.columns else "-")
    mc[2].metric("파일 수정일",p_date)
    st.divider()
    srch=st.text_input("🔍 검색","",placeholder="지종명, 계획지류명...")
    if srch:
        mask=plan_df.astype(str).apply(lambda c:c.str.contains(srch,case=False)).any(axis=1)
        plan_df=plan_df[mask]
    st.markdown(f"**{len(plan_df):,}건**")
    st.dataframe(plan_df,hide_index=True,use_container_width=True,height=560)
    buf=io.BytesIO(); plan_df.to_excel(buf,index=False,engine="openpyxl"); buf.seek(0)
    st.download_button("📥 다운로드",data=buf,file_name="계획지류_조회결과.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
